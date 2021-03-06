from PyQt5.QtWidgets import QFileDialog
from openpyxl import load_workbook
import utils.excelScripts as excelScripts

import functions.tables.fillTable as fillTable

import errors.errors as errors
import errors.success as success
from errors.handleFileReading import handleFileReading

from constants.screens import screens

def jobsTable(self, initCol, initRow, createTable, fillTable, widget):
    
    fileName, _ = QFileDialog.getOpenFileName(self, "Import Jobs Data", "","Excel Files (*.xlsx)")
    
    if not(fileName):
        return 0
    
    wb2 = load_workbook(fileName)
    sheet = wb2.worksheets[0]
    matrix = excelScripts.readMatrix(sheet, initCol, initRow)
    
    #? Create a table that fits the imported file
    createTable.jobTable(self, rows = len(matrix), columns = len(matrix[0]), widget = widget)
    
    #? Fill the table with imported Data
    try:
        fillTable.jobTable(self, matrix, widget = widget)
    except IndexError:
        errors.indexError(self)
    except Exception:
        errors.Error(self)
    else:
        success.importSuccess(self)
        
    #? Check if the imported table is readable
    handleFileReading(self,  widget = widget, screen = screens["JobsInput"])
        
def delayTable(self, initCol, initRow, widget):

    fileName, _ = QFileDialog.getOpenFileName(self, "Import Delay Data", "", "Excel Files (*.xlsx)")
    
    if not(fileName):
        return 0
    
    wb2 = load_workbook(fileName)
    sheet = wb2.worksheets[0]
    seq = excelScripts.readSeq(sheet, initCol, initRow)
    
    #? Fill the table with imported Data
    try:
        fillTable.delayTable(self, seq, widget = widget)
    except IndexError:
        errors.indexError(self)
    except Exception:
        errors.Error(self)
    else:
        success.importSuccess(self)
        
    #? Check if the imported table is readable
    handleFileReading(self,  widget = widget, screen = screens["DelayInput"])           

def preparationTables(self, initCol,initRow, widget):

    fileName, _ = QFileDialog.getOpenFileName(self,"Import Preparation Data", "","Excel Files (*.xlsx)")
    
    if not(fileName):
        return 0
    
    wb2 = load_workbook(fileName)
    widget.prepMatrix=[]
    
    for sheet in wb2.worksheets:
        matrix=excelScripts.readMatrix(sheet,initCol,initRow)
        widget.prepMatrix.append(matrix)
        
    #? Fill the table with imported Data
    try:
        fillTable.preparationTables(self,  widget = widget)
    except IndexError:
        errors.indexError(self)
    except Exception:
        errors.Error(self)
    else:
        success.importSuccess(self)
        
    #? Check if the imported table is readable
    handleFileReading(self,  widget = widget, screen = screens["PreparationInput"])        